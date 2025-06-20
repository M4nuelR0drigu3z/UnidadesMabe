import os
import sys
import logging
import requests
import pandas as pd
from dateutil import parser as dp
import pytz
from datetime import datetime, timedelta
from email.message import EmailMessage
import smtplib
from dotenv import load_dotenv
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.drawing.image import Image
import time

# ----------------------------------------------------------------------
# CARGAR .env
env_path = Path(__file__).parent / ".env"
load_dotenv(dotenv_path=env_path)

# ----------------------------------------------------------------------
# Credenciales y configuración
PHONE_NUMBER_ID   = os.getenv("WHATSAPP_PHONE_NUMBER_ID")
ACCESS_TOKEN      = os.getenv("WHATSAPP_ACCESS_TOKEN")
DESTINOS = [d.strip() for d in os.getenv("WHATSAPP_DESTINOS", "").split(",") if d.strip()]

SAMSARA_API_TOKEN = os.getenv("SAMSARA_API_TOKEN")

SMTP_HOST         = os.getenv("SMTP_HOST", "smtp.example.com")
SMTP_PORT         = int(os.getenv("SMTP_PORT", 587))
SMTP_USER         = os.getenv("SMTP_USER")
SMTP_PASSWORD     = os.getenv("SMTP_PASSWORD")

TEMPLATE_NAME     = os.getenv("TEMPLATE_NAME", "reporte")
LANG_CODE         = os.getenv("LANG_CODE", "es_MX")
MX_TZ             = "America/Mexico_City"

# ----------------------------------------------------------------------
# Logging
logging.basicConfig(
    filename="reporte_logs.log",
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s"
)

# ----------------------------------------------------------------------
# WhatsApp setup
BASE_URL = f"https://graph.facebook.com/v17.0/{PHONE_NUMBER_ID}"
HEADERS  = {"Authorization": f"Bearer {ACCESS_TOKEN}"}

# ----------------------------------------------------------------------
def crear_link_samsara(asset_id, address, lat, lon, dest_name, link_name, description=None, minutos_expiracion=30):
    url = "https://api.samsara.com/live-shares"
    SAMSARA_API_TOKEN = os.getenv("SAMSARA_API_TOKEN")
    headers = {
        "accept": "application/json",
        "authorization": f"Bearer {SAMSARA_API_TOKEN}",
        "content-type": "application/json"
    }
    expires_at = (datetime.utcnow() + timedelta(minutes=minutos_expiracion)).strftime("%Y-%m-%dT%H:%M:%SZ")
    payload = {
        "type": "assetsLocation",
        "assetsLocationLinkConfig": {
            "assetId": str(asset_id),
            "location": {
                "formattedAddress": address,
                "latitude": float(lat),
                "longitude": float(lon),
                "name": dest_name
            }
        },
        "name": link_name,
        "expiresAtTime": expires_at
    }
    if description:
        payload["description"] = description

    response = requests.post(url, json=payload, headers=headers)
    response.raise_for_status()
    resp_json = response.json()
    # <-- ESTA ES LA CORRECCIÓN
    data = resp_json.get("data", {})
    return data.get("liveSharingUrl", ""), data.get("expiresAtTime", expires_at)


def subir_media(path: str) -> str:
    with open(path, "rb") as f:
        files = {
            "file": (
                os.path.basename(path),
                f,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
            "messaging_product": (None, "whatsapp")
        }
        r = requests.post(f"{BASE_URL}/media", headers=HEADERS, files=files)
        r.raise_for_status()
        media_id = r.json()["id"]
        logging.info(f"Media subido, media_id: {media_id}")
        return media_id

def enviar_template(media_id: str, to: str, excel_path: str):
    payload = {
        "messaging_product": "whatsapp",
        "to": to,
        "type": "template",
        "template": {
            "name": TEMPLATE_NAME,
            "language": {"code": LANG_CODE},
            "components": [
                {
                    "type": "header",
                    "parameters": [
                        {
                            "type": "document",
                            "document": {
                                "id": media_id,
                                "filename": os.path.basename(excel_path)
                            }
                        }
                    ]
                }
            ]
        }
    }
    headers = {**HEADERS, "Content-Type": "application/json"}
    r = requests.post(f"{BASE_URL}/messages", json=payload, headers=headers)
    r.raise_for_status()
    msg_id = r.json()["messages"][0]["id"]
    logging.info(f"Template enviado a {to}, message ID: {msg_id}")

def main():
    logging.info("===> Inicio de ejecución")

    if not SAMSARA_API_TOKEN:
        logging.error("❌ Falta SAMSARA_API_TOKEN en variables de entorno")
        sys.exit(1)

    base_dir       = Path(__file__).parent
    plantilla_path = base_dir / "PlantillaMabe.xlsx"

    # IDs predefinidas
    # predefined_special = {
    #     "254792506", "254801835", "254802588",
    #     "254803338", "254859196", "94193861", "95243156",
    #     "95243200", "95243316", "95243513", "244349505",
    #     "245970120", "254794170", "254794716", "257477773"
    # }

    # Obtener nuevos IDs desde Samsara…
    try:
        samsara_h = {
            "Accept": "application/json",
            "Authorization": f"Bearer {SAMSARA_API_TOKEN}"
        }
        tags = requests.get(
            "https://api.samsara.com/tags/4984761",
            headers=samsara_h, timeout=60
        )
        tags.raise_for_status()
        new_ids = {
            a["id"] for a in tags.json()
            .get("data", {}).get("addresses", []) if a.get("id")
        }
    except Exception:
        logging.exception("Error al obtener tags")
        sys.exit(1)

    # Obtener datos GPS…
    try:
        veh = requests.get(
            "https://api.samsara.com/fleet/vehicles/stats?types=gps",
            headers=samsara_h,
            params={"ParentTagIds": "5021095"},
            timeout=60
        )
        veh.raise_for_status()
        vehicles = veh.json().get("data", [])
    except Exception:
        logging.exception("Error al obtener datos de vehículos")
        sys.exit(1)

    results = []
    now_mx = datetime.now(pytz.timezone(MX_TZ))
    
    excluded_ids = {
        "281474987900844",  # DC02
        "281474995920010",  # DC03
        "281474994747511",  # DC05
        "281474995920032",  # DC08

    }
    
    for u in vehicles:
        try:
            if str(u.get("id")) in excluded_ids:
                continue
            gps   = u.get("gps", {})
            t     = gps.get("time")
            if t:
                loc_time = dp.parse(t).astimezone(pytz.timezone(MX_TZ))
                if now_mx - loc_time > timedelta(hours=1):
                    continue
            if gps.get("address", {}).get("id") in new_ids:
                continue
            speed = gps.get("speedMilesPerHour", 0)
            ecu   = gps.get("isEcuSpeed", False)
            # if speed == 0 and not ecu:
            #     continue
            status = "DETENIDO" if (speed == 0 and ecu) else "RUTA"
            location = gps.get("reverseGeo", {}).get("formattedLocation")
            lat = gps.get('latitude')
            lon = gps.get('longitude')
            lat_long = f"{lat},{lon}"
            asset_id = u.get("id", "")
            link = ""
            expira = ""
            try:
                if asset_id and lat and lon and location:
                    link, expira = crear_link_samsara(
                        asset_id=asset_id,
                        address=location,
                        lat=lat,
                        lon=lon,
                        dest_name=location,
                        link_name=u.get("name", "Sin nombre"),
                        minutos_expiracion=30
                    )
                    # Si tienes muchas unidades, pausa aquí para evitar rate limit:
                    time.sleep(1)
            except Exception as ex:
                logging.warning(f"Error creando link para {asset_id}: {ex}")
            results.append({
                "Unidad":    u.get("name", "Sin nombre"),
                "Ubicación": location,
                "Estatus":   status,
                "Coordenadas": lat_long,
                "Link de seguimiento": link,
                "Expira": expira
            })
        except Exception:
            logging.exception(f"Procesando unidad {u.get('name')}")

    # 6) Generar Excel desde plantilla
    wb = load_workbook(filename=plantilla_path)
    ws = wb.active

    # 6.1) Deshacer merges en filas >= 7
    start_row = 8
    for mr in list(ws.merged_cells.ranges):
        if mr.min_row >= start_row:
            ws.unmerge_cells(str(mr))

    # 6.2) Fecha y hora
    ws["C3"] = now_mx.date().isoformat()
    ws["F3"] = now_mx.strftime("%H:%M:%S")

    # 6.3) Preparar estilos
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    green_font = Font(color="006100")   # verde oscuro
    red_fill   = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    red_font   = Font(color="9C0006")    # rojo oscuro

    # 6.4) Volcar datos, fusionar Ubicación C–H, aplicar estilo y bordes
    center_al = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    for i, row in enumerate(results, start=start_row):
        # Unidad
        cell_u = ws.cell(row=i, column=1, value=row["Unidad"])
        cell_u.border = border
        cell_u.alignment = center_al
        
        # Estatus
        cell_s = ws.cell(row=i, column=2, value=row["Estatus"])
        cell_s.border = border
        cell_s.alignment = center_al
        if row["Estatus"] == "DETENIDO":
            cell_s.fill = red_fill
            cell_s.font = red_font
        else:
            cell_s.fill = green_fill
            cell_s.font = green_font

        # Ubicación
        ws.merge_cells(start_row=i, start_column=3, end_row=i, end_column=9)
        for col in range(3, 10):
            cell_loc = ws.cell(row=i, column=col, value=row["Ubicación"] if col==3 else None)
            cell_loc.border = border
            cell_loc.alignment = center_al
        # Coordenadas
        cell_c = ws.cell(row=i, column=10, value=row["Coordenadas"])
        cell_c.border = border 
        cell_c.alignment = center_al
        # Link de seguimiento
        cell_link = ws.cell(row=i, column=11, value=row["Link de seguimiento"])
        cell_link.border = border
        cell_link.alignment = center_al
        # # Expira
        # cell_expira = ws.cell(row=i, column=12, value=row["Expira"])
        # cell_expira.border = border
        # cell_expira.alignment = center_al

    # 6.5) Conteo dinámico en H2
    last_row = start_row + len(results) - 1
    ws["H3"] = f"=COUNTA(A{start_row}:A{last_row})"

    # 6.6) Guardar archivo nuevo
    ws.column_dimensions['A'].width = 15  # ~110px (ajusta si es necesario)
    ws.row_dimensions[2].height = 15     # 45pt = ~60px
    ws.row_dimensions[3].height = 15
    ws.row_dimensions[4].height = 15

    # Crear la imagen y ajustar tamaño
    img_path     = base_dir / "logoMabe.png"
    logo = Image(img_path)
    logo.width = 125    # Ancho total que ocupará (ajusta si el logo es más chico/largo)
    logo.height = 60   # Suma del alto de las 3 filas (45*3)
    ws.add_image(logo, "A2")

    # 6.6) Guardar archivo nuevo
    ts_str        = now_mx.strftime("%Y-%m-%d_%H-%M-%S")
    nuevo_archivo = base_dir / f"Reporte de estatus de unidades {ts_str}.xlsx"
    wb.save(nuevo_archivo)
    logging.info(f"Excel generado: {nuevo_archivo}")

    # 7) Enviar por correo
    try:
        msg = EmailMessage()
        msg['From']    = SMTP_USER
        msg['To'] = "mrodriguez@bgcapitalgroup.mx"
        msg['Subject'] = "Reporte de estatus de unidades"
        msg.set_content("Hola, se adjunta el reporte.\n\nSaludos.")
        with open(nuevo_archivo, "rb") as f:
            msg.add_attachment(
                f.read(),
                maintype="application",
                subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                filename=nuevo_archivo.name
            )
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=60) as s:
            s.starttls()
            s.login(SMTP_USER, SMTP_PASSWORD)
            s.send_message(msg)
        logging.info("Correo enviado")
    except Exception:
        logging.exception("Error enviando correo")
        sys.exit(1)
        
    # 8) Enviar por WhatsApp
    for destino in DESTINOS:
        try:
            media_id = subir_media(str(nuevo_archivo))
            enviar_template(media_id, destino, str(nuevo_archivo))
        except Exception:
            logging.exception(f"Error enviando WhatsApp a {destino}")
            sys.exit(1)
    # 9) Eliminar archivo
    try:
        os.remove(str(nuevo_archivo))
        logging.info(f"Archivo eliminado: {nuevo_archivo.name}")
    except Exception:
        logging.exception(f"No se pudo eliminar: {nuevo_archivo.name}")

    logging.info("===> Ejecución finalizada correctamente")

if __name__ == "__main__":
    main()
